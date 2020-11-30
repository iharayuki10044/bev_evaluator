#!usr/bin/env python2

import sys, os
import math
import numpy as np
import tf
import cv2
from cv_bridge import CvBridge
import rospy
from sensor_msgs.msg import Image
from std_msgs.msg import Float32MultiArray
from geometry_msgs.msg import Pose2D

from skimage.measure import compare_ssim as ssim
from skimage.measure import compare_psnr as psnr

import openpyxl

class BevCalculator:
    def __init__(self):
        rospy.init_node('bev_calculator', anonymous=True)

        # param
        self.HZ = rospy.get_param("~HZ", 10)
        self.MANUAL_CROP_SIZE = rospy.get_param("~MANUAL_CROP_SIZE", 5)
        self.excel_file_name = rospy.get_param("EXCEL_FILE_NAME", "/records/records.xlsx")

        self.input_estimate_img = None
        self.input_flow_img = None
        self.output_true_img = None
        self.current_yaw = None

        self.Bridge = CvBridge()

        # subscriber
        true_img_sub = rospy.Subscriber('/bev_true/true_flow_image',Image ,self.true_img_callback)
        flow_img_sub = rospy.Subscriber('/bev/flow_image' ,Image ,self.flow_img_callback)
        current_pose2d_sub = rospy.Subscriber('/bev_true/current_yaw', Pose2D, self.current_pose2d_callback)

        # publisher
        self.measurement_val_pub = rospy.Publisher('/bev_evaluator/mesurement_val', Float32MultiArray, queue_size = 10)
        self.true_flow_img_pub = rospy.Publisher('/bev_evaluator/true_flow_image', Image, queue_size = 10)

    def true_img_callback(self, data):
        self.input_flow_img = self.Bridge.imgmsg_to_cv2(data, "bgr8")

    def flow_img_callback(self, data):
        self.input_estimate_img = self.Bridge.imgmsg_to_cv2(data, "bgr8")

    def current_pose2d_callback(self, data):
        self.current_yaw = data.theta *180 /math.pi
        while self.current_yaw < 0.0:
            self.current_yaw += 360.0

    def ssim_measurement(self, true_img, flow_img):
        return measurement(ssim, true_img, flow_img)

    def psnr_measurement(self, true_img, flow_img):
        return measurement(psnr, true_img, flow_img)

    def image_rotator(self, input_img, angle):
        height = input_img.shape[0]
        width = input_img.shape[1]
        center = (int(width/2), int(height/2))

        trans = cv2.getRotationMatrix2D(center, angle, 1)
        output_img = cv2.warpAffine(input_img, trans, (width, height))

        return output_img

    def image_cropper(self, input_img, manual_crop_size):
        height = input_img.shape[0]
        width = input_img.shape[1]
        cropped_img = input_img[manual_crop_size:height-manual_crop_size, manual_crop_size:width-manual_crop_size, :]
        return cropped_img

    def process(self):
        r = rospy.Rate(self.HZ)
        # book = openpyxl.load_workbook(filename=self.excel_file_name)
        # i = 0
        while not rospy.is_shutdown():
            
            if self.input_flow_img is not None:
                print("img rotate")
                rotation_true_img = self.image_rotator(self.input_flow_img, self.current_yaw)
                self.output_true_img = self.image_cropper(rotation_true_img, self.MANUAL_CROP_SIZE)
                img_msg = self.Bridge.cv2_to_imgmsg(self.output_true_img, "bgr8")
                self.true_flow_img_pub.publish(img_msg)
            
            measurement_val = [0.0, 0.0]
            if self.input_estimate_img is not None:
                measurement_val[1] = self.ssim_measurement(self.output_true_img, self.input_estimate_img)
                measurement_val[2] = self.psnr_measurement(self.output_true_img, self.input_estimate_img)
                # sheet.cell(row=i, column=1).value = measurement_val[1]
                # sheet.cell(row=i, column=2).value = measurement_val[2]
                # i += 1
                self.measurement_val_pub.publish(2, measurement_val)
                print("result")
                print(measurement_val)
            r.sleep()


if __name__ == '__main__':
    bc = BevCalculator()
    try:
        bc.process()
    except rospy.ROSInterruptException:
        pass
